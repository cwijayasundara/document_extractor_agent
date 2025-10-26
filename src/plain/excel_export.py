"""Excel export utility for extracted data.

This module provides functionality to export extracted document data to Excel format
with proper formatting and multiple sheets for different data types.
"""

from io import BytesIO
from datetime import datetime
from typing import Any
import json

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def flatten_dict(d: dict, parent_key: str = "", sep: str = " → ") -> dict:
    """Flatten a nested dictionary.

    Args:
        d: Dictionary to flatten
        parent_key: Parent key for nested items
        sep: Separator for nested keys

    Returns:
        Flattened dictionary with concatenated keys
    """
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict) and v:
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            if not v:
                items.append((new_key, ""))
            elif all(isinstance(item, dict) for item in v):
                # Skip lists of objects (these are line items)
                continue
            else:
                # Flatten simple lists to comma-separated string
                items.append((new_key, ", ".join(str(item) for item in v if item not in (None, ""))))
        else:
            items.append((new_key, v))
    return dict(items)


def format_value(val: Any) -> Any:
    """Format a value for Excel display.

    Args:
        val: Value to format

    Returns:
        Formatted value suitable for Excel
    """
    if val is None or val == "":
        return ""
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        return val.strip()
    if isinstance(val, list):
        if not val:
            return ""
        return ", ".join(str(item) for item in val if item not in (None, ""))
    return str(val)


def auto_adjust_column_width(sheet, min_width: int = 10, max_width: int = 50):
    """Auto-adjust column widths based on content.

    Args:
        sheet: Worksheet to adjust
        min_width: Minimum column width
        max_width: Maximum column width
    """
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            except:
                pass

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        sheet.column_dimensions[column_letter].width = adjusted_width


def export_to_excel(data: dict, filename: str = "extracted_data.xlsx") -> BytesIO:
    """Export extracted data to Excel format.

    Creates a workbook with three sheets:
    1. Document Headers: Flattened key-value pairs for document-level fields
    2. Line Items: Tabular data for line items (if present)
    3. Raw Data: JSON representation of the full data

    Args:
        data: Extracted data dictionary
        filename: Suggested filename (not used, but kept for API consistency)

    Returns:
        BytesIO object containing the Excel file
    """
    # Detect line items using common field names
    LINE_ITEM_NAMES = ["items", "line_items", "lines", "products", "details", "entries", "lineItems"]
    line_items_field = None
    line_items_data = None
    document_data = {}

    # First pass: look for common line item field names
    for key, value in data.items():
        if key.lower() in [name.lower() for name in LINE_ITEM_NAMES]:
            if isinstance(value, list) and value:
                line_items_field = key
                line_items_data = value
        else:
            document_data[key] = value

    # Second pass: if no common name found, look for any array of objects
    if line_items_field is None:
        for key, value in data.items():
            if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                line_items_field = key
                line_items_data = value
                # Remove from document data since it's line items
                if key in document_data:
                    del document_data[key]
                break

    # Create workbook
    wb = Workbook()

    # Define header style
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_alignment = Alignment(horizontal="left", vertical="center")

    # Sheet 1: Document Headers
    ws_headers = wb.active
    ws_headers.title = "Document Headers"

    # Add headers
    ws_headers["A1"] = "Field"
    ws_headers["B1"] = "Value"
    ws_headers["A1"].font = header_font
    ws_headers["B1"].font = header_font
    ws_headers["A1"].fill = header_fill
    ws_headers["B1"].fill = header_fill
    ws_headers["A1"].alignment = header_alignment
    ws_headers["B1"].alignment = header_alignment

    # Freeze top row
    ws_headers.freeze_panes = "A2"

    # Add document header data
    flattened_data = flatten_dict(document_data)
    row_idx = 2
    for field_name, field_value in sorted(flattened_data.items()):
        ws_headers[f"A{row_idx}"] = field_name
        ws_headers[f"B{row_idx}"] = format_value(field_value)
        row_idx += 1

    # Auto-adjust column widths
    auto_adjust_column_width(ws_headers)

    # Sheet 2: Line Items (if present)
    if line_items_data and len(line_items_data) > 0:
        ws_items = wb.create_sheet("Line Items")

        # Get all unique keys from all line items
        all_keys = set()
        for item in line_items_data:
            if isinstance(item, dict):
                all_keys.update(item.keys())

        # Sort keys for consistent ordering
        sorted_keys = sorted(all_keys)

        # Add headers
        for col_idx, key in enumerate(sorted_keys, start=1):
            cell = ws_items.cell(row=1, column=col_idx)
            # Prettify the column name
            pretty_name = key.replace("_", " ").title()
            cell.value = pretty_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Freeze top row
        ws_items.freeze_panes = "A2"

        # Add data rows
        for row_idx, item in enumerate(line_items_data, start=2):
            if isinstance(item, dict):
                for col_idx, key in enumerate(sorted_keys, start=1):
                    value = item.get(key)
                    ws_items.cell(row=row_idx, column=col_idx).value = format_value(value)

        # Auto-adjust column widths
        auto_adjust_column_width(ws_items)

    # Sheet 3: Raw Data (JSON)
    ws_raw = wb.create_sheet("Raw Data")

    # Add header
    ws_raw["A1"] = "Raw JSON Data"
    ws_raw["A1"].font = header_font
    ws_raw["A1"].fill = header_fill

    # Add formatted JSON
    json_str = json.dumps(data, indent=2)
    ws_raw["A2"] = json_str
    ws_raw["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    # Adjust column width for raw data
    ws_raw.column_dimensions["A"].width = 80

    # Add metadata sheet
    ws_meta = wb.create_sheet("Metadata", 0)  # Insert at beginning
    ws_meta["A1"] = "Export Information"
    ws_meta["A1"].font = Font(bold=True, size=14)

    ws_meta["A3"] = "Export Date:"
    ws_meta["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws_meta["A4"] = "Document Fields:"
    ws_meta["B4"] = len(flattened_data)

    ws_meta["A5"] = "Line Items:"
    ws_meta["B5"] = len(line_items_data) if line_items_data else 0

    if line_items_field:
        ws_meta["A6"] = "Line Items Field:"
        ws_meta["B6"] = line_items_field

    ws_meta["A3"].font = Font(bold=True)
    ws_meta["A4"].font = Font(bold=True)
    ws_meta["A5"].font = Font(bold=True)
    ws_meta["A6"].font = Font(bold=True)

    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 30

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
