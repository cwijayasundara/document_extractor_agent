"""Excel template-based JSON schema generator for research_2.

This module provides simplified template schema generation without AI classification.
Uses simple "Item" prefix heuristic to classify document vs line item fields.
"""

from typing import Dict, Any, List
from pathlib import Path
from openpyxl import load_workbook


class TemplateSchemaGenerator:
    """Generate JSON Schema from Excel template for structured data extraction."""

    def __init__(self, template_path: str):
        """Initialize with Excel template path.

        Args:
            template_path: Path to the Excel template file (.xlsx)

        Raises:
            FileNotFoundError: If template file doesn't exist
            ValueError: If file is not .xlsx format
        """
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        if not str(self.template_path).endswith('.xlsx'):
            raise ValueError(f"Template must be .xlsx format, got: {template_path}")

    def _clean_field_name(self, field_name: str) -> str:
        """Clean Excel header to create valid JSON property name.

        Args:
            field_name: Raw Excel header name

        Returns:
            Cleaned property name (e.g., "*Invoice No" → "Invoice_No")
        """
        if not field_name:
            return ""

        # Remove asterisks (required field markers)
        cleaned = field_name.replace("*", "").strip()

        # Replace spaces and special chars with underscores
        cleaned = cleaned.replace(" ", "_").replace("(", "_").replace(")", "_")
        cleaned = cleaned.replace("/", "_or_").replace("-", "_")

        # Remove consecutive underscores and trailing underscores
        while "__" in cleaned:
            cleaned = cleaned.replace("__", "_")
        cleaned = cleaned.strip("_")

        return cleaned

    def _infer_field_type(self, field_name: str) -> str:
        """Infer JSON type from field name conventions.

        Args:
            field_name: Field name

        Returns:
            JSON Schema type (string, number, or boolean)
        """
        field_lower = field_name.lower()

        # Numeric fields
        if any(keyword in field_lower for keyword in [
            "quantity", "qty", "rate", "amount", "price", "charge",
            "total", "subtotal", "tax", "discount", "cost"
        ]):
            return "number"

        # Boolean fields
        if any(keyword in field_lower for keyword in [
            "taxable", "enabled", "active", "is_", "has_"
        ]):
            return "boolean"

        # Default to string (includes dates, text, etc.)
        return "string"

    def _is_required_field(self, field_name: str) -> bool:
        """Check if field is marked as required (prefixed with *).

        Args:
            field_name: Field name from Excel

        Returns:
            True if field starts with *, False otherwise
        """
        return field_name.strip().startswith("*")

    def get_headers(self) -> List[str]:
        """Extract raw headers from Excel template.

        Returns:
            List of Excel header names from first row

        Raises:
            ValueError: If template has no headers
        """
        wb = load_workbook(self.template_path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1] if cell.value]

        if not headers:
            raise ValueError("Template must have headers in the first row")

        return headers

    def classify_fields(self, headers: List[str]) -> Dict[str, List[str]]:
        """Classify headers into document-level and line-item fields.

        Uses simple heuristic: fields starting with "Item" are line items.

        Args:
            headers: List of Excel header names

        Returns:
            Dictionary with 'document_fields' and 'line_item_fields'
        """
        document_fields = []
        line_item_fields = []

        for header in headers:
            if not header:
                continue

            # Remove asterisk for classification
            clean_header = header.replace("*", "").strip()

            # Fields starting with "Item" are line-item fields
            if clean_header.startswith("Item"):
                line_item_fields.append(header)
            else:
                document_fields.append(header)

        return {
            "document_fields": document_fields,
            "line_item_fields": line_item_fields
        }

    def generate_schema(
        self,
        document_type: str = "document",
        include_line_items: bool = True
    ) -> Dict[str, Any]:
        """Generate JSON Schema from Excel template.

        Args:
            document_type: Type of document (used in description)
            include_line_items: Whether to include line items array

        Returns:
            JSON Schema Draft 7 dictionary

        Raises:
            ValueError: If template has no headers
        """
        # Extract headers from first row
        headers = self.get_headers()

        # Classify fields
        classification = self.classify_fields(headers)
        document_fields = classification["document_fields"]
        line_item_fields = classification["line_item_fields"]

        # Build schema properties
        properties = {}
        required_fields = []

        # Add document-level fields
        for field in document_fields:
            clean_name = self._clean_field_name(field)
            if not clean_name:
                continue

            field_type = self._infer_field_type(field)

            properties[clean_name] = {
                "type": field_type,
                "description": f"{field.replace('*', '').strip()} from {document_type}"
            }

            if self._is_required_field(field):
                required_fields.append(clean_name)

        # Add line items array if requested
        if include_line_items and line_item_fields:
            line_item_properties = {}
            line_item_required = []

            for field in line_item_fields:
                # Remove "Item" prefix for cleaner property names
                field_without_prefix = field.replace("Item", "", 1).strip()
                clean_name = self._clean_field_name(field_without_prefix)

                if not clean_name:
                    continue

                field_type = self._infer_field_type(field)

                line_item_properties[clean_name] = {
                    "type": field_type,
                    "description": f"{field_without_prefix.replace('*', '').strip()} for this line item"
                }

                if self._is_required_field(field):
                    line_item_required.append(clean_name)

            # Add line_items array to properties
            properties["line_items"] = {
                "type": "array",
                "description": f"List of line items from {document_type}",
                "items": {
                    "type": "object",
                    "properties": line_item_properties,
                    "required": line_item_required if line_item_required else []
                }
            }

        # Build complete schema
        schema = {
            "type": "object",
            "description": f"Structured data extracted from {document_type}",
            "properties": properties,
            "required": required_fields if required_fields else []
        }

        return schema

    def get_field_summary(self) -> Dict[str, Any]:
        """Get summary of template fields for debugging/display.

        Returns:
            Dictionary with field classification and counts
        """
        headers = self.get_headers()
        classification = self.classify_fields(headers)

        return {
            "total_fields": len(headers),
            "document_fields": classification["document_fields"],
            "line_item_fields": classification["line_item_fields"],
            "document_field_count": len(classification["document_fields"]),
            "line_item_field_count": len(classification["line_item_fields"]),
        }


def generate_schema_from_template(
    template_path: str,
    document_type: str = "document"
) -> Dict[str, Any]:
    """Convenience function to generate schema from Excel template.

    Args:
        template_path: Path to Excel template (.xlsx)
        document_type: Type of document (for description)

    Returns:
        JSON Schema dictionary

    Raises:
        FileNotFoundError: If template doesn't exist
        ValueError: If template is invalid
    """
    generator = TemplateSchemaGenerator(template_path)
    return generator.generate_schema(document_type=document_type)
